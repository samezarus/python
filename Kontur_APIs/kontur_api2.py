#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# pip3 install httplib2
# pip3 install openpyxl
# pip3 install pymongo
# pip3 install pymysql

import httplib2, urllib
import json
import openpyxl
import pymongo
import pymysql
from datetime import datetime

class TKkt():
    # из json-a
    regNumber      = '' # Регистрациооный номер (РНМ кассы)
    name           = '' # имя ККТ
    serialNumber   = '' # Заводской номер ККТ
    organizationId = '' # id организации
    modelName      = '' # Название модели
    salesPointId   = '' # id точки продаж
    fnSerialNumber = '' # Заводской номер ФН
    # из xlsx "Список касс в Excel"
    ffd          = '' # "ФФД"
    state        = '' # "Состояние ККТ"
    lastDoc      = '' # "Последний документ"
    lastDocDate  = '' # "Дата последнего документа"
    endPay       = '' # "Оплата" (Срок окончания оплаты ОФД)
    ofd          = '' # "ОФД"
    fnChangeDate = '' # "Срок замены ФН"
    # из эксельки Райгородского
    firmware = '' # Версия прошивки ККТ
    mac      = '' # MAC - АДРЕС
    memo     = '' # Примечание
    # из УКМ-а
    sgoName = '' # имя c СГО сервера
    #
    addDate = '' # Дата создания/обновления записи

class TKktStatistics():
    # Текущее состояние на кассе с Регистрационным номером "cashboxRegNumber"
    cashboxRegNumber = ''
    #
    buy_cashlessTotal = 0.0
    buy_cashTotal = 0.0
    buy_total = 0.0
    buy_revenue = 0.0
    buy_count = 0
    #
    returnBuy_cashlessTotal = 0.0
    returnBuy_cashTotal = 0.0
    returnBuy_total = 0.0
    returnBuy_revenue = 0.0
    returnBuy_count = 0
    # --- чеки продаж
    sell_cashlessTotal = 0.0  # Безналичными
    sell_cashTotal = 0.0  # Наличными
    sell_total = 0.0  # Выручка
    sell_revenue = 0.0
    sell_count = 0  # Количество чеков
    # --- чеки возвратов
    returnSell_cashlessTotal = 0.0
    returnSell_cashTotal = 0.0
    returnSell_total = 0.0
    returnSell_revenue = 0.0
    returnSell_count = 0

class TSalesPoint():
    organizationId = '' # id Организации за которой закреплена данная точка продаж
    id             = '' # id Точки продажи
    name           = '' # Имя Точки продажи
    sgoName        = '' # имя c СГО сервера
    #
    addDate = '' # Дата создания/обновления записи

class TOrganization():
    id             = '' # id организации в Контур
    inn            = '' # ИНН организации
    kpp            = '' # КПП организации
    shortName      = '' # Краткое наименование организации
    fullName       = '' # Полное наименование организации
    isBlocked      = '' # Заблокирована ли организация или нет
    creationDate   = '' #
    hasEvotorOffer = '' #
    #
    addDate = ''  # Дата создания/обновления записи

class TKontur():
    ofd_api_key  = '' # ключ интегратора
    SID          = '' #
    cookies      = {}
    cookieDomain = '' #
    endPoint     = '' # площадка
    mail         = '' # емаил авторизации
    passwd       = '' # пароль авторизации

    authStatusCode = '' # код статуса авторизации

    Organizations = [] # Список организаций
    SalesPoints   = [] # Списка точек продаж организации
    Kkts          = [] # Список ККТ
    KktStatistics = [] # Список текущих статусов продаж касс

    sgoServer   = ''
    sgoUser     = ''
    sgoPassword = ''
    sgoDb       = ''

    lastError = ''

    addonXlsxFile = '' # экселька Райгородского



    def __init__(self):
        conn = pymongo.MongoClient('localhost', 27017)
        db   = conn['TKontur']

        self.dbOrgs          = db['orgs']
        self.dbSps           = db['sps']
        self.dbKkts          = db['kkts']
        self.dbKktStatistics = db['kktStatistics']

    def end_pay_format(self, s):
        # "28.07.2019 - 28.08.2020" to "28.08.2020"
        result = s
        if len(s) == 23:
            if s[2] == '.' and s[5] == '.' and s[11] == '-' and s[15] == '.' and s[18] == '.':
                result = s[13:23]

        return result

    def nowRuDate(self):
        s = str(datetime.now())
        s = s[8:10] + '.' + s[5:7] + '.' + s[0:4]
        return s

    def authenticate_by_pass(self):
        # Аутентификация/авторизация и получение SID-a (https://kontur-ofd-api.readthedocs.io/ru/latest/Auth/authenticate-by-pass.html)
        url_str             = 'https://api.kontur.ru/auth/authenticate-by-pass?login=' + self.mail  # урл авторизации
        h                   = httplib2.Http('.cache')
        (resp, content)     = h.request(uri=url_str, method='POST', body=self.passwd)
        self.authStatusCode = str(resp.status)

        if self.authStatusCode == '200':
            self.SID = content.decode("utf-8")
            self.SID = self.SID[8:]
            self.SID = self.SID[:-2]

            self.cookies = {'Cookie': 'ofd_api_key=' + self.ofd_api_key + ';auth.sid=' + self.SID}
        else:
            self.lastError = 'Не могу авторизоваться'

    def get_organizations(self):
        # метод organizations (https://kontur-ofd-api.readthedocs.io/ru/latest/http/organizations.html)
        if self.authStatusCode == '200':
            url_str         = self.endPoint + 'v1/organizations'
            h               = httplib2.Http('.cache')
            (resp, content) = h.request(uri=url_str, method='GET', headers=self.cookies)

            if str(resp.status) == '200':
                json_obj = json.loads(content.decode("utf-8"))
                for item in json_obj:
                    Organization                = TOrganization()
                    Organization.id             = item ['id']
                    Organization.inn            = item ['inn']
                    #Organization.kpp            = item ['kpp'] # нет у ИП
                    Organization.shortName      = item ['shortName']
                    Organization.fullName       = item ['fullName']
                    Organization.isBlocked      = item ['isBlocked']
                    #Organization.creationDate   = item ['creationDate'] # нет у пустой организации
                    #Organization.hasEvotorOffer = item ['hasEvotorOffer']
                    Organization.addDate = datetime.now()

                    self.Organizations.append (Organization)
            else:
                self.lastError = 'Не могу получить список организаций'

    def get_sales_point(self):
        # метод salespoints (недокументированный)
        # получение списка точек продаж организации
        for org in self.Organizations:
            url_str         = self.endPoint + 'v1/organizations/' + org.id + '/salesPoints'
            h               = httplib2.Http('.cache')
            (resp, content) = h.request(uri=url_str, method='GET', headers=self.cookies)

            if str(resp.status) == '200':
                json_obj = json.loads(content.decode("utf-8"))
                for item in json_obj:
                    SalesPoint                = TSalesPoint()
                    SalesPoint.organizationId = item['organizationId']
                    SalesPoint.id             = item['id']
                    SalesPoint.name           = item['name']
                    SalesPoint.addDate = datetime.now()

                    self.SalesPoints.append(SalesPoint)
            else:
                self.lastError = 'Не могу получить список точек продаж у ' + org.shortName

    def get_cashboxes(self):
        # метод cashboxes (https://kontur-ofd-api.readthedocs.io/ru/latest/http/cashboxes.html)
        # получаем списки ККТ по всем организациям доступным пользователю
        for org in self.Organizations:
            url_str         = self.endPoint + 'v1/organizations/' + org.id + '/cashboxes'
            h               = httplib2.Http('.cache')
            (resp, content) = h.request(uri=url_str, method='GET', headers=self.cookies)

            if str(resp.status) == '200':
                json_obj = json.loads(content.decode("utf-8"))

                for item in json_obj:
                    Kkt                = TKkt()
                    Kkt.regNumber      = item['regNumber']
                    Kkt.name           = item['name']
                    Kkt.serialNumber   = item['serialNumber']
                    Kkt.organizationId = item['organizationId']
                    Kkt.modelName      = item['modelName']
                    Kkt.salesPointId   = item['salesPointPeriods'][0]['salesPointId'] # !!! поидеи надо в цикле дёргать
                    Kkt.fnSerialNumber = item['fnEntity']['serialNumber']
                    Kkt.addDate = datetime.now()

                    self.Kkts.append(Kkt)
            else:
                self.lastError = 'Не могу получить список ККТ у ' + org.shortName

    def get_sp_tickets(self, orgId, spId, fromDate, toDate):
        # Получение сумарной информации по каждой кассе конкретной Организации на конкретной Точке продаж

        # сумарно по вусем точкам https://ofd-api.kontur.ru/v1/organizations/16c9ac54-76a6-4610-9b4a-4306f8b7edba/statistics/cashReceipt?From=01.10.2019&To=01.10.2019
        # по каждой точке https://ofd-api.kontur.ru/v1/organizations/16c9ac54-76a6-4610-9b4a-4306f8b7edba/statistics/cashReceipt/salesPoints?From=01.10.2019&To=01.10.2019
        # по сумарно конкретной точке https://ofd-api.kontur.ru/v1/organizations/16c9ac54-76a6-4610-9b4a-4306f8b7edba/statistics/cashReceipt/salesPoints/7e563e8d-4d67-450f-a60e-1561c315f9a9?From=01.10.2019&To=01.10.2019
        # список чеков по конкретной точке по каждой кассе https://ofd-api.kontur.ru/v1/organizations/16c9ac54-76a6-4610-9b4a-4306f8b7edba/statistics/cashReceipt/salesPoints/7e563e8d-4d67-450f-a60e-1561c315f9a9/cashboxes?From=01.10.2019&To=01.10.2019

        url_str = self.endPoint + 'v1/organizations/'+orgId+'/statistics/cashReceipt/salesPoints/'+spId+'/cashboxes?From='+fromDate+'&To='+toDate

        h = httplib2.Http('.cache')
        (resp, content) = h.request(uri=url_str, method='GET', headers=self.cookies)

        if str(resp.status) == '200':
            json_obj = json.loads(content.decode("utf-8"))

            for item in json_obj:
                #print (item['cashboxRegNumber'])
                for kkt in self.Kkts:
                    if kkt.regNumber == item['cashboxRegNumber']:
                        #print (item['items'][0]['sell']['total'])

                        KktStatistic = TKktStatistics()
                        #
                        KktStatistic.cashboxRegNumber = item['cashboxRegNumber']
                        #
                        KktStatistic.buy_cashlessTotal = item['items'][0]['buy']['cashlessTotal']
                        KktStatistic.buy_cashTotal     = item['items'][0]['buy']['cashTotal']
                        KktStatistic.buy_total         = item['items'][0]['buy']['total']
                        KktStatistic.buy_revenue       = item['items'][0]['buy']['revenue']
                        KktStatistic.buy_count         = item['items'][0]['buy']['count']
                        #
                        KktStatistic.returnBuy_cashlessTotal = item['items'][0]['returnBuy']['cashlessTotal']
                        KktStatistic.returnBuy_cashTotal     = item['items'][0]['returnBuy']['cashTotal']
                        KktStatistic.returnBuy_total         = item['items'][0]['returnBuy']['total']
                        KktStatistic.returnBuy_revenue       = item['items'][0]['returnBuy']['revenue']
                        KktStatistic.returnBuy_count         = item['items'][0]['returnBuy']['count']
                        #
                        KktStatistic.sell_cashlessTotal = item['items'][0]['sell']['cashlessTotal']
                        KktStatistic.sell_cashTotal     = item['items'][0]['sell']['cashTotal']
                        KktStatistic.sell_total         = item['items'][0]['sell']['total']
                        KktStatistic.sell_revenue       = item['items'][0]['sell']['revenue']
                        KktStatistic.sell_count         = item['items'][0]['sell']['count']
                        #
                        KktStatistic.returnSell_cashlessTotal = item['items'][0]['returnSell']['cashlessTotal']
                        KktStatistic.returnSell_cashTotal     = item['items'][0]['returnSell']['cashTotal']
                        KktStatistic.returnSell_total         = item['items'][0]['returnSell']['total']
                        KktStatistic.returnSell_revenue       = item['items'][0]['returnSell']['revenue']
                        KktStatistic.returnSell_count         = item['items'][0]['returnSell']['count']

                        self.KktStatistics.append(KktStatistic)

    def get_statistics(self):
        for org in self.Organizations:
            for sp in self.SalesPoints:
                if sp.organizationId == org.id:
                    self.get_sp_tickets(org.id, sp.id, self.nowRuDate(), self.nowRuDate())


    def get_cashboxes_xlsx(self):
        for org in self.Organizations:
            url_str = self.endPoint + 'v1/organizations/' + org.id + '/cashbox-connection/cashboxes'
            h = httplib2.Http('.cache')
            (resp, content) = h.request(uri=url_str, method='GET', headers=self.cookies)

            if str(resp.status) == '200':
                xlsx = 'log.xlsx'
                f = open(xlsx, 'wb') #
                f.write(content)
                f.close()

                wb = openpyxl.load_workbook(xlsx)
                ws = wb.worksheets[0]

                for i in range(2, ws.max_row+1):
                    for kkt in self.Kkts:
                        if kkt.serialNumber == str(ws.cell(row=i, column=1).value):
                            kkt.ffd          = str(ws.cell(row=i, column=6).value)
                            kkt.state        = str(ws.cell(row=i, column=7).value)
                            kkt.lastDoc      = str(ws.cell(row=i, column=8).value)
                            kkt.lastDocDate  = str(ws.cell(row=i, column=9).value)
                            kkt.endPay       = self.end_pay_format(str(ws.cell(row=i, column=10).value))
                            kkt.ofd          = str(ws.cell(row=i, column=14).value)
                            kkt.fnChangeDate = str(ws.cell(row=i, column=15).value)

                            break
            else:
                self.lastError = 'Не могу получить файл эксель с доп. инф. по ККТ у ' + org.shortName

    def addon_xlsx(self):
        wb = openpyxl.load_workbook(self.addonXlsxFile)
        ws = wb.worksheets[0]

        for i in range(2, ws.max_row + 1):
            for kkt in self.Kkts:
                if kkt.serialNumber == str(ws.cell(row=i, column=10).value):
                    kkt.firmware = str(ws.cell(row=i, column=13).value)
                    kkt.mac      = str(ws.cell(row=i, column=15).value)
                    kkt.memo     = str(ws.cell(row=i, column=19).value)

                    break

    def org_to_mongo(self):
        for org in self.Organizations:
            doc = {"id": str(org.id),
                   "inn": str(org.inn),
                   "kpp": str(org.kpp),
                   "shortName": str(org.shortName),
                   "fullName": str(org.fullName),
                   "isBlocked": str(org.isBlocked),
                   "creationDate": str(org.creationDate),
                   "hasEvotorOffer": str(org.hasEvotorOffer),
                   "addDate": str(org.addDate),
                   }

            if self.dbOrgs.find({"id" : str(org.id)}).count() == 0:
                self.dbOrgs.save(doc)
            else:
                self.dbOrgs.update({"id": str(org.id)}, doc)

    def sp_to_mongo(self):
        for sp in self.SalesPoints:
            doc = {"organizationId": str(sp.organizationId),
                   "id": str(sp.id),
                   "name": str(sp.name),
                   "addDate": str(sp.addDate),
                   }

            if self.dbSps.find({"id": str(sp.id)}).count() == 0:
                self.dbSps.save(doc)
            else:
                self.dbSps.update({"id": str(sp.id)}, doc)

    def kkt_to_mongo(self):
        for kkt in self.Kkts:
            doc = {"regNumber": str(kkt.regNumber),
                   "name": str(kkt.name),
                   "serialNumber": str(kkt.serialNumber),
                   "organizationId": str(kkt.organizationId),
                   "modelName": str(kkt.modelName),
                   "salesPointId": str(kkt.salesPointId),
                   "fnSerialNumber": str(kkt.fnSerialNumber),
                   "ffd": str(kkt.ffd),
                   "state": str(kkt.state),
                   "lastDoc": str(kkt.lastDoc),
                   "lastDocDate": str(kkt.lastDocDate),
                   "endPay": str(kkt.endPay),
                   "ofd": str(kkt.ofd),
                   "fnChangeDate": str(kkt.fnChangeDate),
                   "firmware": str(kkt.firmware),
                   "mac": str(kkt.mac),
                   "memo": str(kkt.memo),
                   "sgoName": str(kkt.sgoName),
                   "addDate": str(kkt.addDate),
                   }

            if self.dbKkts.find({"serialNumber": str(kkt.serialNumber)}).count() == 0:
                self.dbKkts.save(doc)
            else:
                self.dbKkts.update({"serialNumber": str(kkt.serialNumber)}, doc)

    def kktStatistic_to_mongo(self):
        for KktStatistic in self.KktStatistics:
            doc = {"cashboxRegNumber": KktStatistic.cashboxRegNumber,
                   "buy_cashlessTotal": KktStatistic.buy_cashlessTotal,
                   "buy_total": KktStatistic.buy_total,
                   "buy_revenue": KktStatistic.buy_revenue,
                   "buy_count": KktStatistic.buy_count,
                   "returnBuy_cashlessTotal": KktStatistic.returnBuy_cashlessTotal,
                   "returnBuy_cashTotal": KktStatistic.returnBuy_cashTotal,
                   "returnBuy_total": KktStatistic.returnBuy_total,
                   "returnBuy_revenue": KktStatistic.returnBuy_revenue,
                   "returnBuy_count": KktStatistic.returnBuy_count,
                   "sell_cashlessTotal": KktStatistic.sell_cashlessTotal,
                   "sell_cashTotal": KktStatistic.sell_cashTotal,
                   "sell_total": KktStatistic.sell_total,
                   "sell_revenue": KktStatistic.sell_revenue,
                   "sell_count": KktStatistic.sell_count,
                   "returnSell_cashlessTotal": KktStatistic.returnSell_cashlessTotal,
                   "returnSell_cashTotal": KktStatistic.returnSell_cashTotal,
                   "returnSell_total": KktStatistic.returnSell_total,
                   "returnSell_revenue": KktStatistic.returnSell_revenue,
                   "returnSell_count": KktStatistic.returnSell_count,
                   }

            if self.dbKktStatistics.find({"cashboxRegNumber": str(KktStatistic.cashboxRegNumber)}).count() == 0:
                self.dbKktStatistics.save(doc)
            else:
                self.dbKktStatistics.update({"cashboxRegNumber": str(KktStatistic.cashboxRegNumber)}, doc)


    def from_sgo(self):
        #
        connection = pymysql.connect(host=self.sgoServer,
                                     user=self.sgoUser,
                                     password=self.sgoPassword,
                                     db=self.sgoDb,
                                     charset='utf8mb4',
                                     cursorclass=pymysql.cursors.DictCursor
                                     )
        try:
            conn   = pymongo.MongoClient('localhost', 27017)
            db     = conn['TKontur']
            dbUkms = db['ukms']

            with connection.cursor() as cursor:
                # Дополняеи инфу о Точках продаж из СГО

                # Дополняеи инфу о ККТ из СГО
                sql = 'SELECT DISTINCT trm_out_shift_close.kkm_serial_number, trm_in_pos.name from trm_out_shift_close, trm_in_pos WHERE trm_out_shift_close.cash_id = trm_in_pos.cash_id'
                cursor.execute(sql)
                for row in cursor:
                    for kkt in self.Kkts:
                        if row['kkm_serial_number'] ==  kkt.serialNumber:
                            kkt.sgoName = row['name']
                            break
        finally:
            connection.close()
# --------------------------------------------------------------------------------------------------------------------

Kontur = TKontur()

#

Kontur.ofd_api_key   = '' # ключ для доступа личного кабинета через браузер
Kontur.SID           = ''
Kontur.cookieDomain  = '.kontur.ru'
Kontur.endPoint      = 'https://ofd-api.kontur.ru/'
Kontur.mail          = ''
Kontur.passwd        = ''
Kontur.sgoServer     = ''
Kontur.sgoUser       = ''
Kontur.sgoPassword   = ''
Kontur.sgoDb         = ''
Kontur.addonXlsxFile = 'addon.xlsx'

#

Kontur.authenticate_by_pass()
#print (Kontur.SID)
#print (Kontur.lastError)
#

Kontur.get_organizations()
#print (len(Kontur.Organizations))
#print (Kontur.lastError)
#

Kontur.get_sales_point()
#print (Kontur.lastError)
#

Kontur.get_cashboxes()
#print (Kontur.lastError)
#

Kontur.get_cashboxes_xlsx()
#print (Kontur.lastError)
#

Kontur.get_statistics()
#print (Kontur.lastError)
#

Kontur.addon_xlsx()

#

Kontur.from_sgo()

#

Kontur.org_to_mongo()
Kontur.sp_to_mongo()
Kontur.kkt_to_mongo()
Kontur.kktStatistic_to_mongo()

#

"""
f = open('log.html', 'w')
f.write('<html>')
f.write('<body>')

for org in Kontur.Organizations:
    print (org.shortName)
    f.write('<br>')
    f.write('<p><font color="#59b300"><h1>'+org.shortName+'</h1></font></p>')
    for sp in Kontur.SalesPoints:
        if sp.organizationId == org.id:
            print ('  '+sp.name)
            f.write('<h3>' + sp.name + '</h3>')

            f.write('<table border="1" cellpadding="5" cellspacing="0">')
            f.write('<tr>')
            f.write('<th>Название ККТ</th>')
            f.write('<th>Заводской №</th>')
            f.write('<th>Регистрационный №</th>')
            f.write('<th>Заводской № ФН</th>')
            f.write('<th>ФФД</th>')
            f.write('<th>Состояние ККТ</th>')
            f.write('<th>Последний документ</th>')
            f.write('<th>Дата последнего документа</th>')
            f.write('<th>Оплата</th>')
            f.write('<th>ОФД</th>')
            f.write('<th>Срок замены ФН</th>')
            f.write('</tr>')

            for kkt in Kontur.Kkts:
                if (kkt.organizationId == org.id) and (kkt.salesPointId == sp.id):
                    f.write('<tr>')
                    print ('    ' + kkt.regNumber)
                    print ('    ' + kkt.endPay)
                    f.write('<td>' + kkt.name + '</td>')
                    f.write('<td>' + kkt.serialNumber + '</td>')
                    f.write('<td>' + kkt.regNumber + '</td>')
                    f.write('<td>' + kkt.fnSerialNumber + '</td>')

                    f.write('<td>' + kkt.ffd + '</td>')
                    f.write('<td>' + kkt.state + '</td>')
                    f.write('<td>' + kkt.lastDoc  + '</td>')
                    f.write('<td>' + kkt.lastDocDate + '</td>')
                    f.write('<td>' + kkt.endPay + '</td>')
                    f.write('<td>' + kkt.ofd  + '</td>')
                    f.write('<td>' + kkt.fnChangeDate + '</td>')
                    f.write('</tr>')

            f.write('</table>')

f.write('</body>')
f.write('</html>')
f.close()

"""
